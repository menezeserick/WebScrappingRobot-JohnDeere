# -*- coding: utf-8 -*-
"""
Processador de Excel com ArcGIS

Este script lê um arquivo Excel contendo as colunas:
    Dealer, Endereco 1, Endereco 2

Para cada linha, monta o endereço completo, consulta o ArcGIS para obter cidade, estado, latitude, longitude, etc. 
Em seguida, gera um novo arquivo Excel com os resultados.

Interface gráfica: o usuário seleciona o arquivo de entrada e o local/arquivo de saída por meio de diálogos.
A interface foi feita com Tkinter, deixando a experiência mais agradável.

Para gerar um EXE:
    1. Instale o PyInstaller: pip install pyinstaller
    2. Gere o executável com o comando:
         pyinstaller --onefile --windowed nome_do_script.py
"""

import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from arcgis.geocoding import geocode
from arcgis.gis import GIS
from openpyxl import Workbook, load_workbook

gis = GIS()

def arcgis_search(endereco_completo):
    """
    Busca informações de geocodificação usando o ArcGIS.
    Retorna um dicionário com os dados ou valores padrão em caso de erro.
    """
    try:
        geo_code = geocode(endereco_completo)[0]
        if geo_code:
            attributes = geo_code.get("attributes", {})
            cidade = attributes.get("City", "Cidade ND")
            estado = attributes.get("Region", "Estado ND")
            
            return {
                "address": geo_code.get('address', 'AddressGis ND'),
                "lat": geo_code['location'].get('y', 'lat ND'),
                "lng": geo_code['location'].get('x', 'lng ND'),
                "postal": geo_code['attributes'].get('Postal', 'Zip ND'),
                "cidade": cidade,
                "estado": estado
            }
    except Exception as e:
        print(f"Erro na busca do ArcGIS: {e}")
    return {
        "address": "AddressGis ND",
        "lat": "lat ND",
        "lng": "lng ND",
        "postal": "Zip ND",
        "cidade": "Cidade ND",
        "estado": "Estado ND"
    }

def process_excel(input_file, output_file, log_callback):
    """
    Processa o arquivo Excel de entrada e salva os resultados em output_file.
    O parâmetro log_callback é uma função que registra as mensagens de log na interface.
    """
    try:
        wb_input = load_workbook(input_file)
        ws_input = wb_input.active

        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = "Resultados"

        headers = [
            "Dealer", "Endereco Completo", "Cidade", "Estado", "Pais",
            "Latitude", "Longitude", "CEP", "Endereco GIS"
        ]
        ws_output.append(headers)

        total_rows = ws_input.max_row - 1
        log_callback(f"Total de linhas para processar: {total_rows}")

        for i, row in enumerate(ws_input.iter_rows(min_row=2, values_only=True), start=1):
            try:
                log_callback(f"Processando linha {i}...")
                dealer = row[0]
                endereco1 = row[1]
                endereco2 = row[2]

                if not all([dealer, endereco1]):
                    log_callback(f"Linha {i} ignorada: Dados insuficientes.")
                    continue

                endereco_completo = f"{endereco1} {endereco2 if endereco2 else ''}, Brasil"

                geo_code = arcgis_search(endereco_completo)

                ws_output.append([
                    dealer, endereco_completo,
                    geo_code["cidade"], geo_code["estado"], "Brasil",
                    geo_code["lat"], geo_code["lng"], geo_code["postal"], geo_code["address"]
                ])
            except Exception as e:
                log_callback(f"Erro ao processar linha {i}: {e}")

        wb_output.save(output_file)
        wb_output.close()
        log_callback(f"Processamento concluído. Resultados salvos em:\n{output_file}")
    except Exception as e:
        log_callback(f"Erro ao processar o arquivo: {e}")

class ExcelProcessorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Processador de Excel ArcGIS")
        self.geometry("600x400")
        self.resizable(False, False)

        self.input_file = tk.StringVar()
        self.output_file = tk.StringVar()

        self.create_widgets()

    def create_widgets(self):
        frame_files = tk.Frame(self)
        frame_files.pack(pady=10, padx=10, fill=tk.X)

        label_input = tk.Label(frame_files, text="Arquivo de Entrada:")
        label_input.grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        entry_input = tk.Entry(frame_files, textvariable=self.input_file, width=50)
        entry_input.grid(row=0, column=1, padx=5, pady=5)
        btn_browse_input = tk.Button(frame_files, text="Selecionar", command=self.select_input_file)
        btn_browse_input.grid(row=0, column=2, padx=5, pady=5)

        label_output = tk.Label(frame_files, text="Arquivo de Saída:")
        label_output.grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        entry_output = tk.Entry(frame_files, textvariable=self.output_file, width=50)
        entry_output.grid(row=1, column=1, padx=5, pady=5)
        btn_browse_output = tk.Button(frame_files, text="Salvar Como", command=self.select_output_file)
        btn_browse_output.grid(row=1, column=2, padx=5, pady=5)

        self.btn_process = tk.Button(
            self,
            text="Processar",
            command=self.start_processing,
            bg="#4CAF50",
            fg="white",
            font=("Helvetica", 12, "bold")
        )
        self.btn_process.pack(pady=10)

        self.log_text = scrolledtext.ScrolledText(
            self, wrap=tk.WORD, width=70, height=10, state="disabled", bg="#F0F0F0"
        )
        self.log_text.pack(pady=10, padx=10)

    def select_input_file(self):
        filename = filedialog.askopenfilename(
            title="Selecione o arquivo Excel de entrada",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filename:
            self.input_file.set(filename)

    def select_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="Selecione onde salvar o arquivo Excel de saída",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filename:
            self.output_file.set(filename)

    def log_message(self, message):
        """
        Insere a mensagem de log na área de texto de forma thread-safe.
        """
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")

    def start_processing(self):
        input_path = self.input_file.get()
        output_path = self.output_file.get()

        if not input_path:
            messagebox.showerror("Erro", "Por favor, selecione o arquivo de entrada.")
            return

        if not os.path.exists(input_path):
            messagebox.showerror("Erro", "O arquivo de entrada não foi encontrado.")
            return

        if not output_path:
            messagebox.showerror("Erro", "Por favor, selecione o local para salvar o arquivo de saída.")
            return

        self.btn_process.config(state="disabled")
        self.log_message("Iniciando o processamento...")

        thread = threading.Thread(target=self.run_processing, args=(input_path, output_path))
        thread.start()

    def run_processing(self, input_path, output_path):
        process_excel(input_path, output_path, self.log_message)
        self.log_message("Processamento finalizado.")
        self.btn_process.config(state="normal")
        messagebox.showinfo("Concluído", "O processamento foi concluído com sucesso!")

if __name__ == "__main__":
    app = ExcelProcessorApp()
    app.mainloop()
